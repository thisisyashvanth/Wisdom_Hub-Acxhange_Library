from openpyxl import Workbook
from io import BytesIO


def generate_users_excel(users: list):

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = ["User ID", "Name", "Employee ID", "Email", "Role", "Restricted"]
    ws.append(headers)

    for user in users:
        ws.append([
            user.get("id"),
            user.get("name"),
            user.get("employee_id"),
            user.get("email"),
            user.get("role"),
            "Restricted" if user.get("is_restricted") else "Active"
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def generate_requests_excel(requests: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Request History"

    headers = [
        "Request ID",
        "Book",
        "Employee",
        "Type",
        "Requested On",
        "Reviewed On",
        "Status",
        "Remarks"
    ]
    ws.append(headers)

    for r in requests:
        ws.append([
            r.get("id"),
            r.get("book_title"),
            r.get("user_name"),
            r.get("request_type"),
            r.get("request_date"),
            r.get("return_date"),
            r.get("status"),
            r.get("remarks")
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def generate_books_excel(books: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Books"

    headers = [
        "Book ID",
        "Title",
        "Book Number",
        "Author",
        "Category",
        "ISBN",
        "Total Copies",
        "Available Copies",
        "Borrowed Copies",
        "Status"
    ]
    ws.append(headers)

    for book in books:
        total_copies = book.get("total_copies") or 0
        available_copies = book.get("available_copies") or 0

        if available_copies == 0:
            status = "Unavailable"
        else:
            status = f"{available_copies} available"

        ws.append([
            book.get("id"),
            book.get("title"),
            book.get("bookNumber"),
            book.get("author"),
            book.get("category"),
            book.get("isbn"),
            total_copies,
            available_copies,
            total_copies - available_copies,
            status
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def generate_book_history_excel(history: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Book History"

    headers = [
        "Borrow ID",
        "User ID",
        "Employee",
        "Employee ID",
        "Issue Date",
        "Due Date",
        "Returned Date",
        "Status",
        "Renewals"
    ]
    ws.append(headers)

    for record in history:
        ws.append([
            record.get("borrow_id"),
            record.get("user_id"),
            record.get("employee_name"),
            record.get("employee_id"),
            record.get("issue_date"),
            record.get("due_date"),
            record.get("returned_date") or "-",
            record.get("status"),
            record.get("renewal_count")
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream